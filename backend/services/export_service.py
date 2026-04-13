"""
export_service.py
─────────────────
Provides CSV and XLSX exports for trays and scan log.
XLSX uses openpyxl (must be installed: pip install openpyxl).
"""
import csv
import io
from datetime import datetime
from sqlalchemy.orm import Session
from models import Tray, ScanEvent


# ── CSV exports ───────────────────────────────────────────────────────────────

def export_trays_csv(
    db:         Session,
    tenant_id:  str            = "default",
    stage:      str | None     = None,
    project:    str | None     = None,
    start_date: datetime | None = None,
    end_date:   datetime | None = None,
) -> bytes:
    """Export trays to CSV bytes. Accepts optional filters."""
    q = db.query(Tray).filter(Tray.tenant_id == tenant_id)

    if stage:      q = q.filter(Tray.stage   == stage)
    if project:    q = q.filter(Tray.project == project)
    if start_date: q = q.filter(Tray.created_at >= start_date)
    if end_date:   q = q.filter(Tray.created_at <= end_date)

    trays = q.order_by(Tray.created_at.desc()).all()

    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow([
        "ID", "Stage", "Project", "Shift", "Batch No", "Created By",
        "Total Units", "FIFO Violated", "Is Done", "Is Split Parent",
        "Parent ID", "Created At", "Last Updated", "Completed At",
    ])
    for t in trays:
        w.writerow([
            t.id, t.stage, t.project, t.shift, t.batch_no or "", t.created_by,
            t.total_units, "YES" if t.fifo_violated else "NO",
            "YES" if t.is_done else "NO",
            "YES" if t.is_split_parent else "NO",
            t.parent_id or "",
            _fmt(t.created_at), _fmt(t.last_updated), _fmt(t.completed_at),
        ])

    return buf.getvalue().encode("utf-8-sig")   # utf-8-sig for Excel compatibility


def export_scan_log_csv(
    db:        Session,
    tenant_id: str = "default",
    limit:     int = 50_000,
) -> bytes:
    """Export scan event log to CSV bytes."""
    events = (
        db.query(ScanEvent)
        .filter(ScanEvent.tenant_id == tenant_id)
        .order_by(ScanEvent.timestamp.desc())
        .limit(limit)
        .all()
    )

    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["Timestamp", "Tray ID", "From Stage", "To Stage",
                "Operator", "FIFO Flag", "Note"])
    for e in events:
        w.writerow([
            _fmt(e.timestamp),
            e.tray_id, e.from_stage, e.stage,
            e.operator,
            "YES" if e.fifo_flag else "NO",
            e.note or "",
        ])

    return buf.getvalue().encode("utf-8-sig")


# ── XLSX export ───────────────────────────────────────────────────────────────

def export_report_xlsx(db: Session, tenant_id: str = "default") -> bytes:
    """Full production report as a multi-sheet XLSX workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for XLSX export. Run: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()

    # ── Sheet 1: Trays ────────────────────────────────────────────────────────
    ws1    = wb.active
    ws1.title = "Trays"

    tray_headers = [
        "ID", "Stage", "Project", "Shift", "Batch No", "Created By",
        "Total Units", "FIFO Violated", "Done", "Split Parent",
        "Parent ID", "Created At", "Last Updated", "Completed At",
    ]
    _write_header_row(ws1, tray_headers, fill_hex="185FA5")

    trays = (
        db.query(Tray)
        .filter(Tray.tenant_id == tenant_id)
        .order_by(Tray.created_at.desc())
        .all()
    )
    for row_i, t in enumerate(trays, 2):
        row_data = [
            t.id, t.stage, t.project, t.shift, t.batch_no or "", t.created_by,
            t.total_units,
            "YES" if t.fifo_violated    else "NO",
            "YES" if t.is_done          else "NO",
            "YES" if t.is_split_parent  else "NO",
            t.parent_id or "",
            _fmt(t.created_at), _fmt(t.last_updated), _fmt(t.completed_at),
        ]
        for col_i, val in enumerate(row_data, 1):
            ws1.cell(row=row_i, column=col_i, value=val)

        # Light red fill for FIFO-violated rows
        if t.fifo_violated:
            for col_i in range(1, len(tray_headers) + 1):
                ws1.cell(row=row_i, column=col_i).fill = PatternFill(
                    "solid", fgColor="FECACA"
                )

    _autofit(ws1, tray_headers)

    # ── Sheet 2: Scan Log ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Scan Log")
    log_headers = ["Timestamp", "Tray ID", "From Stage", "To Stage",
                   "Operator", "FIFO", "Note"]
    _write_header_row(ws2, log_headers, fill_hex="3B6D11")

    events = (
        db.query(ScanEvent)
        .filter(ScanEvent.tenant_id == tenant_id)
        .order_by(ScanEvent.timestamp.desc())
        .limit(50_000)
        .all()
    )
    for row_i, e in enumerate(events, 2):
        ws2.cell(row=row_i, column=1, value=_fmt(e.timestamp))
        ws2.cell(row=row_i, column=2, value=e.tray_id)
        ws2.cell(row=row_i, column=3, value=e.from_stage)
        ws2.cell(row=row_i, column=4, value=e.stage)
        ws2.cell(row=row_i, column=5, value=e.operator)
        ws2.cell(row=row_i, column=6, value="YES" if e.fifo_flag else "NO")
        ws2.cell(row=row_i, column=7, value=e.note or "")

    _autofit(ws2, log_headers)

    # ── Sheet 3: Stage Summary ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Stage Summary")
    _write_header_row(ws3, ["Stage", "Tray Count"], fill_hex="7F77DD")

    stage_counts: dict = {}
    for t in trays:
        stage_counts[t.stage] = stage_counts.get(t.stage, 0) + 1

    for row_i, (stage, count) in enumerate(sorted(stage_counts.items()), 2):
        ws3.cell(row=row_i, column=1, value=stage)
        ws3.cell(row=row_i, column=2, value=count)

    _autofit(ws3, ["Stage", "Tray Count"])

    # ── Serialise to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Worksheet helpers ─────────────────────────────────────────────────────────

def _write_header_row(ws, headers: list, fill_hex: str = "185FA5"):
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_i, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col_i, value=h)
        cell.font       = Font(bold=True, color="FFFFFF")
        cell.fill       = PatternFill("solid", fgColor=fill_hex)
        cell.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _autofit(ws, headers: list, min_w: int = 10, max_w: int = 40):
    for col_i, h in enumerate(headers, 1):
        ws.column_dimensions[
            __import__("openpyxl.utils", fromlist=["get_column_letter"]).get_column_letter(col_i)
        ].width = min(max(len(h) + 4, min_w), max_w)


def _fmt(dt) -> str:
    if not dt:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return str(dt)
